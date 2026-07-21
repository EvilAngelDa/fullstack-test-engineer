#!/usr/bin/env python3
"""Write QA cases to an Excel file matching a common import template.

Input JSON (list or {"cases": [...]}):
{
  "cases": [
    {
      "name": "【展示】模块可见",
      "module": "/App/Module/UI",
      "tags": "前端;功能;P0",
      "precondition": "已登录；进入详情页",
      "steps": [
        {"action": "打开页面", "expected": "模块展示"},
        {"action": "点击展开", "expected": "展示全文"}
      ],
      "mode": "STEP",
      "remark": "",
      "status": "Prepare",
      "owner": "",
      "level": "P0"
    }
  ]
}

Usage:
  python3 write_cases_xlsx.py --out cases.xlsx --cases-json cases.json
  python3 write_cases_xlsx.py --out cases.xlsx --cases-json -  # stdin
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List


def load_cases(path: str) -> List[Dict[str, Any]]:
    if path == "-":
        data = json.load(sys.stdin)
    else:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "cases" in data:
        return data["cases"]
    raise ValueError("JSON must be a list or {\"cases\": [...]}")


def write_xlsx(cases: List[Dict[str, Any]], out: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required: pip install openpyxl\n" + str(e)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        "用例名称",
        "所属模块",
        "标签",
        "前置条件",
        "步骤描述",
        "预期结果",
        "编辑模式",
        "备注",
        "用例状态",
        "责任人",
        "用例等级",
    ]
    widths = [42, 30, 24, 48, 58, 58, 12, 36, 12, 12, 10]
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True, name="Microsoft YaHei", size=11)
    cell_font = Font(name="Microsoft YaHei", size=10)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, i, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w

    comments = {
        2: "若无该模块将自动创建",
        3: "标签之间以分号或者逗号隔开",
        5: "STEP 模式下可多行步骤",
        6: "与步骤一一对应",
        7: "STEP 或 TEXT",
        9: 'Prepare / Underway / Completed',
        10: "项目内人员 ID（可选）",
        11: "P0 / P1 / P2 / P3",
    }
    for col, text in comments.items():
        ws.cell(1, col).comment = Comment(text, "fullstack-test-engineer")

    row = 2
    for case in cases:
        steps = case.get("steps") or []
        if not steps:
            steps = [{"action": case.get("action", ""), "expected": case.get("expected", "")}]
        # normalize
        norm = []
        for s in steps:
            if isinstance(s, (list, tuple)) and len(s) >= 2:
                norm.append({"action": s[0], "expected": s[1]})
            elif isinstance(s, dict):
                norm.append(
                    {
                        "action": s.get("action") or s.get("step") or "",
                        "expected": s.get("expected") or s.get("expect") or "",
                    }
                )
            else:
                norm.append({"action": str(s), "expected": ""})
        steps = norm
        start = row
        end = row + len(steps) - 1
        values = {
            1: case.get("name") or case.get("title") or "",
            2: case.get("module") or "",
            3: case.get("tags") or "",
            4: case.get("precondition") or case.get("pre") or "",
            5: steps[0]["action"],
            6: steps[0]["expected"],
            7: case.get("mode") or "STEP",
            8: case.get("remark") or "",
            9: case.get("status") or "Prepare",
            10: case.get("owner") or "",
            11: case.get("level") or case.get("priority") or "P1",
        }
        for col, val in values.items():
            cell = ws.cell(start, col, val)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = center if col in (7, 9, 11) else wrap

        for i, st in enumerate(steps[1:], 1):
            r = start + i
            for col in range(1, 12):
                c = ws.cell(r, col, "" if col == 3 else None)
                c.border = thin
                c.font = cell_font
                c.alignment = wrap
            for col, val in ((5, st["action"]), (6, st["expected"])):
                c = ws.cell(r, col, val)
                c.font = cell_font
                c.border = thin
                c.alignment = wrap

        if end > start:
            for col in (1, 2, 3, 4, 7, 8, 9, 10, 11):
                ws.merge_cells(
                    start_row=start, start_column=col, end_row=end, end_column=col
                )
                ws.cell(start, col).alignment = (
                    center if col in (7, 9, 11) else wrap
                )
        for r in range(start, end + 1):
            ws.row_dimensions[r].height = 48
        row = end + 1

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:K{row - 1}"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {len(cases)} cases -> {out}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--cases-json", required=True, help="Path or - for stdin")
    args = ap.parse_args()
    cases = load_cases(args.cases_json)
    write_xlsx(cases, Path(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
